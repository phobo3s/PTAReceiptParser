"""
Excel Muhasebe Defteri Güncelleyici
====================================
Fiş parser çıktısını (Receipt + kategorize edilmiş kalemler) Excel double-entry
defterine yazar. Tarih + tutar ile eşleşen transaction'ın to-account satırlarını
kalem kalem günceller.

Excel sütun yapısı:
    A  Tarih           (D.MM.YYYY — örn: 4.12.2025)
    B  Transaction Code
    C  Payee/Note
    D  Notes (kullanılmıyor)
    E  CURRENCY::TRY
    F  Operasyon (BUY/SELL/Interest/Dividend)
    G  Tag/Note
    H  Full Account Name  ← from-account (1. satır) ve to-accounts (sonraki satırlar)
    I  Tutar              ← Türk formatı: nokta=binlik, virgül=ondalık
    J  Rate/Price         (1,0000000)
    K  Reconciliation
    L  Not
    M  Bank Desc.

Transaction yapısı:
    1. satır  → A=tarih, B=txcode, C=payee, E=currency, H=from_hesap, I=tutar(neg)
    n. satır  → A=boş, H=to_hesap, I=tutar(pos), G=ürün_adı
    (sonraki A dolu satıra kadar devam eder)
"""

import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from parser import Receipt, ReceiptItem

AMOUNT_TOLERANCE = 0.02  # TL eşleşme toleransı


# ── Yardımcı: Tutar dönüşümleri ───────────────────────────────────────────────

def parse_excel_amount(value) -> Optional[float]:
    """
    Türk formatındaki tutar stringini float'a çevirir.
    '2.194,32' → 2194.32
    '-194,32'  → -194.32
    Sayısal değer (int/float) olarak gelirse doğrudan döndürür.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    # Türk formatı: nokta binlik ayracı, virgül ondalık
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def format_excel_amount(amount: Optional[float]) -> str:
    """
    Float'ı Türk formatına çevirir.
    2194.32 → '2.194,32'
    -194.0  → '-194,00'
    """
    if amount is None:
        return "0,00"
    # Binlik ayraç olarak nokta, ondalık olarak virgül
    formatted = f"{abs(amount):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if amount < 0:
        formatted = f"-{formatted}"
    return f"{formatted}" 


# ── Yardımcı: Tarih dönüşümleri ───────────────────────────────────────────────

def parse_excel_date(value) -> Optional[str]:
    """
    Excel hücresinden gelen tarihi 'YYYY-MM-DD' formatına çevirir.
    Kabul edilen girişler:
        - datetime / date nesnesi (openpyxl bazen böyle okur)
        - '4.12.2025', '04.12.2025'  (D.MM.YYYY veya DD.MM.YYYY)
    """
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    # D.MM.YYYY veya DD.MM.YYYY
    m = re.match(r"^(\d{1,2})\.(\d{2})\.(\d{4})$", s)
    if m:
        day, month, year = m.group(1), m.group(2), m.group(3)
        return f"{year}-{month}-{int(day):02d}"
    return None


def receipt_date_to_excel(receipt_date: str) -> str:
    """
    'YYYY-MM-DD' → 'D.MM.YYYY' (Excel formatı, baştaki sıfır yok)
    '2025-12-04' → '4.12.2025'
    """
    dt = datetime.strptime(receipt_date, "%Y-%m-%d")
    return f"{dt.day}.{dt.month:02d}.{dt.year}"


# ── Yardımcı: Merged cell güvenli silme/ekleme ────────────────────────────────

def _unmerge_rows(ws, first_row: int, last_row: int) -> None:
    """Verilen satır aralığıyla çakışan tüm merged cell'leri unmerge eder.
    delete_rows / insert_rows öncesi çağrılmazsa openpyxl geçersiz XML üretir."""
    to_unmerge = [
        str(mc)
        for mc in list(ws.merged_cells.ranges)
        if mc.min_row <= last_row and mc.max_row >= first_row
    ]
    for mc_str in to_unmerge:
        ws.unmerge_cells(mc_str)


# ── Excel transaction eşleştirme ──────────────────────────────────────────────

def find_excel_transaction(ws, receipt: Receipt, tolerance: float = AMOUNT_TOLERANCE) -> Optional[int]:
    """
    Worksheet'te Receipt'e uyan from-account satırının satır numarasını döndürür.
    Eşleşme kriterleri:
        - A sütunu: tarih == receipt.date
        - I sütunu: |tutar| ≈ receipt.total (±tolerance)
    Bulunamazsa None döner.
    """
    if not receipt.date or not receipt.total:
        return None

    for row in ws.iter_rows(min_row=1):
        a_cell = row[0]  # A sütunu
        i_cell = row[8]  # I sütunu (0-indexed → 8)

        if a_cell.value is None:
            continue

        row_date = parse_excel_date(a_cell.value)
        if row_date != receipt.date:
            continue

        row_amount = parse_excel_amount(i_cell.value)
        if row_amount is None:
            continue

        if abs(abs(row_amount) - receipt.total) <= tolerance:
            return a_cell.row

    return None


def get_to_account_rows(ws, from_row: int) -> tuple[int, int]:
    """
    from_row'dan sonraki to-account satır aralığını döndürür.
    To-account satırları: A sütunu boş olan ard arda satırlar.
    Döndürür: (first_to_row, last_to_row)
    Eğer hiç to-account satırı yoksa: (from_row + 1, from_row) — boş aralık
    """
    max_row = ws.max_row
    first_to = from_row + 1
    last_to = from_row  # başlangıçta boş aralık

    for r in range(from_row + 1, max_row + 2):
        if r > max_row:
            break
        a_val = ws.cell(row=r, column=1).value
        if a_val is not None and str(a_val).strip():
            break  # yeni transaction başladı
        last_to = r

    return first_to, last_to


# ── Eşleşme kontrolü (process_receipt'ten çağrılır) ──────────────────────────

def find_excel_match(
    excel_path: Path,
    receipt: Receipt,
    sheet_name: Optional[str] = None,
) -> tuple[Optional[int], Optional[str]]:
    """
    Excel dosyasında eşleşen from-account satırını bul.
    Döndürür: (row_number, account_name) — bulunamazsa (None, None).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ❌ openpyxl kurulu değil. Kurmak için: pip install openpyxl")
        return None, None

    if not excel_path.exists():
        print(f"  ❌ Excel dosyası bulunamadı: {excel_path}")
        return None, None

    keep_vba = str(excel_path).lower().endswith(".xlsm")
    wb = load_workbook(str(excel_path), keep_vba=keep_vba)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"  ❌ Sheet bulunamadı: '{sheet_name}'")
            return None, None
        ws = wb[sheet_name]
    else:
        ws = wb.active

    if ws is None:
        return None, None
    
    from_row = find_excel_transaction(ws, receipt)
    if from_row is None:
        return None, None

    account = ws.cell(row=from_row, column=8).value  # H sütunu
    return from_row, str(account) if account else ""


# ── Mevcut to-account okuma ───────────────────────────────────────────────────

def read_excel_to_accounts(
    excel_path: Path,
    receipt: Receipt,
    sheet_name: Optional[str] = None,
) -> dict[float, str]:
    """
    Excel'deki mevcut to-account satırlarını okur.
    Döndürür: {tutar: hesap_adı} — eşleşme bulunamazsa boş dict.
    Birden fazla satır aynı tutara sahipse sonuncusu kazanır.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    if not excel_path.exists():
        return {}

    keep_vba = str(excel_path).lower().endswith(".xlsm")
    wb = load_workbook(str(excel_path), keep_vba=keep_vba, read_only=True)
    if sheet_name:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    else:
        ws = wb.active
    if ws is None:
        return {}

    from_row = find_excel_transaction(ws, receipt)
    if from_row is None:
        return {}

    first_to, last_to = get_to_account_rows(ws, from_row)
    if last_to < first_to:
        return {}

    result: dict[float, str] = {}
    for r in range(first_to, last_to + 1):
        account = ws.cell(row=r, column=8).value  # H
        amount  = parse_excel_amount(ws.cell(row=r, column=9).value)  # I
        if account and amount is not None:
            result[round(amount, 2)] = str(account)

    return result


# ── Önizleme ──────────────────────────────────────────────────────────────────

def preview_excel(categorized: list[tuple[ReceiptItem, str]], receipt: Receipt):
    print("\n" + "═" * 60)
    print("  Önizleme — Excel'e yazılacak to-account satırları:")
    print("═" * 60)
    print(f"  Fiş: {receipt.store}  {receipt.date}  {format_excel_amount(receipt.total)} TRY")
    print()
    for item, account in categorized:
        amount_str = format_excel_amount(item.amount)
        print(f"  H: {account:<45}  I: {amount_str:>12}  G: {item.name}")
    print(f"\n  Toplam: {format_excel_amount(receipt.total)}")
    print("═" * 60)


# ── Ana güncelleme fonksiyonu ─────────────────────────────────────────────────

def _apply_to_ws(ws, receipt: Receipt, categorized: list[tuple[ReceiptItem, str]]) -> Optional[int]:
    """Tek bir fişi açık worksheet'e yazar. Kaydetmez.
    Başarılıysa from_row (int) döner, bulunamazsa None."""
    from_row = find_excel_transaction(ws, receipt)
    if from_row is None:
        print(f"  ❌ Excel'de eşleşen satır bulunamadı!")
        total_str = f"{receipt.total:.2f} TL" if receipt.total else "toplam bilinmiyor"
        print(f"     Aranan: {receipt.date}  {total_str}")
        return None

    print(f"  ✓ Eşleşen satır: {from_row}  ({ws.cell(row=from_row, column=1).value}  {ws.cell(row=from_row, column=8).value})")

    first_to, last_to = get_to_account_rows(ws, from_row)
    existing_to_count = max(0, last_to - first_to + 1)

    # Mevcut to-account'ları silmeden önce D sütununa yedekle
    if existing_to_count > 0:
        from datetime import date as _date
        parts = []
        for r in range(first_to, last_to + 1):
            acc = ws.cell(row=r, column=8).value  # H
            from_acc = ws.cell(row=from_row, column=8).value  # H (from-account)
            if acc and acc != from_acc:
                parts.append(str(acc))
        if parts:
            ws.cell(row=from_row, column=4).value = f"[{_date.today()}] " + " | ".join(parts)

        _unmerge_rows(ws, first_to, last_to)
        ws.delete_rows(first_to, existing_to_count)

    new_count = len(categorized)
    if new_count > 0:
        _unmerge_rows(ws, first_to, first_to)
        ws.insert_rows(first_to, new_count)

    for idx, (item, account) in enumerate(categorized):
        r = first_to + idx
        ws.cell(row=r, column=8).value = account
        cell_i = ws.cell(row=r, column=9)
        cell_i.value = round(item.amount, 2)
        cell_i.number_format = '#,##0.00'
        ws.cell(row=r, column=10).value = "1"
        ws.cell(row=r, column=7).value = item.name

    return from_row


def update_excel(
    excel_path: Path,
    receipt: Receipt,
    categorized: list[tuple[ReceiptItem, str]],
    sheet_name: Optional[str] = None,
) -> bool:
    """Tek fiş için: aç → yaz → kaydet."""
    return update_excel_batch(excel_path, [(receipt, categorized)], sheet_name)


def update_excel_batch(
    excel_path: Path,
    items: list[tuple[Receipt, list[tuple[ReceiptItem, str]]]],
    sheet_name: Optional[str] = None,
) -> list[Optional[int]]:
    """
    Birden fazla fişi tek load+save ile yazar.
    Döndürür: her fiş için from_row (başarılıysa int, bulunamazsa None).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ❌ openpyxl kurulu değil. Kurmak için: pip install openpyxl")
        return [None] * len(items)

    if not excel_path.exists():
        print(f"  ❌ Excel dosyası bulunamadı: {excel_path}")
        return [None] * len(items)

    keep_vba = str(excel_path).lower().endswith(".xlsm")
    wb = load_workbook(str(excel_path), keep_vba=keep_vba)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"  ❌ Sheet bulunamadı: '{sheet_name}'")
            print(f"     Mevcut sheetler: {', '.join(wb.sheetnames)}")
            return [None] * len(items)
        ws = wb[sheet_name]
    else:
        ws = wb.active
    if ws is None:
        return [None] * len(items)

    results = [_apply_to_ws(ws, receipt, categorized) for receipt, categorized in items]

    wb.save(str(excel_path))
    return results
